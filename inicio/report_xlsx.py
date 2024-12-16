from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image
import os
from django.conf import settings
from django.http import HttpResponse

celdas = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']

def insert_header_image(sheet):
    """
    Inserta una imagen de encabezado en la hoja del libro de Excel.
    """
    # Ruta de la imagen
    imagen_path = os.path.join(settings.BASE_DIR, 'inicio', 'static', 'img', 'header_image_uts.jpg')

    # Verifica si la imagen existe
    if not os.path.exists(imagen_path):
        raise FileNotFoundError(f"La imagen no existe en la ruta: {imagen_path}")

    # Inserta la imagen en el Excel
    img = Image(imagen_path)
    img.width = 700
    img.height = 100
    sheet.add_image(img, "B2")  # Inserta la imagen en la celda B2

def reporte_info(sheet, data):
    # Unión de celdas
    sheet.merge_cells('A11:H11')
    sheet['A11'].font = Font(color = 'FFFFFF', bold=True, size=12)
    sheet['A11'].fill = PatternFill('solid', start_color="d20606")
    sheet['A11'] = 'CONTROL DE REPORTES MENSUALES DE SERVICIOS: BIBLIOTECA.'

    sheet.merge_cells('A12:H12')
    sheet['A12'].font = Font(color = '000000', bold=True, size=12)
    sheet['A12'].fill = PatternFill('solid', start_color="d3c905")
    sheet['A12'] = 'CONSULTAS  EN EL CICLO DEL MES DE: ' + data['ciclo']

# Crea la tabla de acervo
def table_acervo(sheet, data):
    # Crea titulo de tabla
    # Unión de celdas
    sheet.merge_cells('A11:H11')
    sheet['A11'].font = Font(color = '000000', bold=True, size=12)
    sheet['A11'].fill = PatternFill('solid', start_color="d3c905")
    sheet['A11'] = 'REPORTE GENERAL DE ACERVO BIBLIOGRÁFICO: ' + data['ciclo']

    # Crea encabezados de la tabla
    sheet['A12'] = "No."
    sheet['A12'].font = Font(color = '000000', bold=True, size=12)
    for c in celdas:
        sheet[c + '12'].fill = PatternFill('solid', start_color="68685f")
        sheet[c + '13'].fill = PatternFill('solid', start_color="68685f")
    sheet.column_dimensions['A'].width = 16
    sheet['B12'] = 'Área de conociento'



def create_excel(data):
    """
    Crea un archivo Excel con una imagen de encabezado.
    """
    # Crear el libro y la hoja
    book = Workbook()
    sheet = book.active

    # Inserta la imagen en el encabezado
    insert_header_image(sheet)
    # Inserta información
    table_acervo(sheet, data)

    return book

def generate_report(data):
    try:
        # Crear el archivo Excel
        book = create_excel(data)

        # Configurar la respuesta HTTP para la descarga
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="Reporte mensual '+ data['ciclo'] +'.xlsx"'

        # Guardar el archivo Excel directamente en la respuesta
        book.save(response)

        return response
    except FileNotFoundError as e:
        # Manejo específico para errores relacionados con la imagen
        return HttpResponse(str(e), status=404)
    except Exception as e:
        # Manejo genérico de errores
        return HttpResponse(f"Error al generar el reporte: {str(e)}", status=500)
